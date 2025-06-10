# =============================================================================
# Script: Raspagem de dados SAU + Atualização em Planilha Excel
# Autor: Alexandre Hesse
# =============================================================================

# ================================ IMPORTAÇÕES ================================
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import openpyxl
import pandas as pd
import time
import tkinter as tk
from tkinter import ttk, messagebox
from ttkbootstrap import Style

from datetime import datetime
from step2 import calcula_tempo
from step3 import etapa_final

# ================================ CONFIGURAÇÕES ================================
arq_default = "./Default.xlsx"
relatorio_inicial = './relatorios/etapa1.xlsx'
tecnicos = ['alexandreh']
# tecnicos = ['alexandreh', 'gabrielca', 'josejn', 'rafaelvb', 'robsonb', 'walterl', 'washingtonc', 'williamp']

# ================================ FUNÇÃO PRINCIPAL ================================
def relatorio_etapa1():

    # --------------------- Função: Limpar Combobox ---------------------
    def limpar():
        cbx_mes_inicio.set('')
        cbx_ano_inicio.set('')
        cbx_mes_termino.set('')
        cbx_ano_termino.set('')

    # --------------------- Função: Atualiza DF com os dados raspados ---------------------
    def atualizar_dados(df, ocorrencia, motivo, data_abertura, data_inicio):
        df['Motivo'] = df['Motivo'].astype(str)
        df['Data_Abertura'] = df['Data_Abertura'].astype(str)
        df['Data_Inicio'] = df['Data_Inicio'].astype(str)

        df.loc[df['Ocorrencia'] == ocorrencia, ['Motivo', 'Data_Abertura', 'Data_Inicio']] = [
            motivo, data_abertura, data_inicio
        ]

    # --------------------- Função: Inserir as datas no DataFrame ---------------------
    def inserir_datas():
        df = pd.read_excel(relatorio_inicial)

        # Adiciona colunas caso não existam
        for col in ['Motivo', 'Data_Abertura', 'Data_Inicio']:
            if col not in df.columns:
                df[col] = ''

        try:
            ocorrencias = df['Ocorrencia'].tolist()
        except KeyError:
            print("Erro: A coluna 'Ocorrencia' não foi encontrada no arquivo Excel.")
            return

        ocorrencias_falhadas = []

        # Inicia navegador
        with webdriver.Chrome() as driver:
            driver.set_window_size(700, 700)
            driver.get("https://servico.intra.pg/sau/login.asp")

            # Login
            driver.find_element(By.NAME, 'txtLogin').send_keys("alexandreh")
            driver.find_element(By.NAME, 'txtSenha').send_keys("230636")
            driver.find_element(By.NAME, 'button').click()

            # Raspagem das ocorrências
            for i in ocorrencias:
                try:
                    driver.get(f'https://servico.intra.pg/sau/ocorr_detalhes.asp?cd={i}')
                    wait = WebDriverWait(driver, 5)

                    motivo = wait.until(EC.presence_of_element_located((By.XPATH, 
                        '/html/body/form[1]/div[3]/table/tbody/tr/td/table/tbody/tr[9]/td[2]/b'))).text
                    data_abertura = wait.until(EC.presence_of_element_located((By.XPATH, 
                        '/html/body/form[1]/div[3]/table/tbody/tr/td/table/tbody/tr[11]/td[2]'))).text
                    data_inicio = wait.until(EC.presence_of_element_located((By.XPATH, 
                        '/html/body/form[1]/div[3]/table/tbody/tr/td/table/tbody/tr[12]/td[2]'))).text

                    atualizar_dados(df, i, motivo, data_abertura, data_inicio)
                except Exception:
                    ocorrencias_falhadas.append(i)

            # Segunda tentativa para falhas
            for i in ocorrencias_falhadas:
                try:
                    driver.get(f'https://servico.intra.pg/sau/ocorr_detalhes.asp?cd={i}')
                    wait = WebDriverWait(driver, 2)

                    motivo = wait.until(EC.presence_of_element_located((By.XPATH, 
                        '/html/body/form[1]/div[3]/table/tbody/tr/td/table/tbody/tr[7]/td[2]/b'))).text
                    data_abertura = wait.until(EC.presence_of_element_located((By.XPATH, 
                        '/html/body/form[1]/div[3]/table/tbody/tr/td/table/tbody/tr[9]/td[2]'))).text
                    data_inicio = wait.until(EC.presence_of_element_located((By.XPATH, 
                        '/html/body/form[1]/div[3]/table/tbody/tr/td/table/tbody/tr[10]/td[2]'))).text
                    


                    atualizar_dados(df, i, motivo, data_abertura, data_inicio)
                except Exception as e:
                    print(f"Erro ao processar a ocorrência {i}: {e}")

            df.to_excel(relatorio_inicial, index=False)
            # messagebox.showwarning("Alerta", "Tarefa Concluída")
            janela.destroy()

            time.sleep(1)
            calcula_tempo()

    # --------------------- Função: Inicia pesquisa e salva no Excel ---------------------
    def iniciar_pesquisa():
        mes_ini = cbx_mes_inicio.get()
        ano_ini = cbx_ano_inicio.get()
        mes_fim = cbx_mes_termino.get()
        ano_fim = cbx_ano_termino.get()

        driver = webdriver.Chrome()
        workbook = openpyxl.load_workbook(arq_default)
        sheet = workbook.active
        proxima_linha = sheet.max_row + 1 if sheet.max_row > 1 else 2

        for tecnico in tecnicos:
            driver.get(
                f'https://servico.intra.pg/sau/relatorios/ocorrenciasportecnico.asp?cboTecnico={tecnico}&cboMesInicio={mes_ini}&cboAnoInicio={ano_ini}&cboMesFim={mes_fim}&cboAnoFim={ano_fim}&Button=OK')
            
            tabela = driver.find_element(By.TAG_NAME, 'table')
            linhas = tabela.find_elements(By.TAG_NAME, 'tr')[2:-1]

            for linha in linhas:
                dados_linha = [tecnico] + [''] * (sheet.max_column - 1)
                colunas = linha.find_elements(By.TAG_NAME, 'td')

                for j, coluna in enumerate(colunas, 1):
                    strong = coluna.find_elements(By.TAG_NAME, 'strong')
                    texto = coluna.text.replace(strong[0].text, '') if strong else coluna.text
                    dados_linha[j] = texto.strip()

                for k, valor in enumerate(dados_linha, 1):
                    sheet.cell(row=proxima_linha, column=k).value = valor
                proxima_linha += 1

        driver.quit()
        workbook.save(relatorio_inicial)
        inserir_datas()



        # ================================ INTERFACE TKINTER ================================
    style = Style(theme='solar')
    style.configure('.', font=('Helvetica', 10))
    janela = style.master
    janela.title("Relatórios - Pesquisa Rápida")
    janela.resizable(False, False)

    # Centraliza janela na tela
    largura_janela, altura_janela = 390, 260
    pos_x = (janela.winfo_screenwidth() - largura_janela) // 2
    pos_y = (janela.winfo_screenheight() - altura_janela) // 2
    janela.geometry(f"{largura_janela}x{altura_janela}+{pos_x}+{pos_y}")

    # Combobox valores
    anos = [str(i) for i in range(2021, 2041)]
    meses = [str(i) for i in range(1, 13)]

    # Layout Interface
    ttk.Label(janela, text='', width=10).grid(row=0, column=0)
    ttk.Label(janela, text='', width=10).grid(row=0, column=1)
    ttk.Label(janela, text='', width=10).grid(row=0, column=2)
    ttk.Label(janela, text='', width=10).grid(row=0, column=3)

    ttk.Label(janela, font=('Helvetica', 8), text='Mês:').grid(row=2, column=2, padx=20, sticky='w')
    ttk.Label(janela, font=('Helvetica', 8), text='Ano:').grid(row=2, column=3, padx=20, sticky='w')
    ttk.Label(janela, font=('Helvetica', 11), text='Início:').grid(row=3, column=1, padx=20, pady=10, sticky='e')
    ttk.Label(janela, font=('Helvetica', 11), text='Término:').grid(row=4, column=1, padx=20, pady=10, sticky='e')

    cbx_mes_inicio = ttk.Combobox(janela, values=meses, width=3)
    cbx_ano_inicio = ttk.Combobox(janela, values=anos, width=5)
    cbx_mes_termino = ttk.Combobox(janela, values=meses, width=3)
    cbx_ano_termino = ttk.Combobox(janela, values=anos, width=5)

    cbx_mes_inicio.grid(row=3, column=2, padx=20, pady=10, sticky='we')
    cbx_ano_inicio.grid(row=3, column=3, padx=20, pady=10, sticky='we')
    cbx_mes_termino.grid(row=4, column=2, padx=20, pady=10, sticky='we')
    cbx_ano_termino.grid(row=4, column=3, padx=20, pady=10, sticky='we')

    ttk.Button(janela, text='Pesquisar', bootstyle="info", command=iniciar_pesquisa).grid(row=5, column=0, columnspan=2, padx=20, pady=10, sticky='we')
    ttk.Button(janela, text='Limpar', command=limpar).grid(row=5, column=2, columnspan=2, padx=20, pady=10, sticky='we')

    ttk.Label(janela, text='Dev: Alexandre Hesse', font=('Arial', 7)).grid(row=7, column=2, columnspan=2, padx=20, pady=10, sticky='e')

    janela.mainloop()

    
    messagebox.showwarning("Alerta", "Tarefa Concluída")

# ================================ EXECUTA ================================

if __name__ == "__main__":
    relatorio_etapa1()


